from django.core import paginator
from django.forms.renderers import get_template
from django.shortcuts import render, redirect
from .models import *
from .forms import *


from django.shortcuts import get_object_or_404

from django.http import HttpResponse
#WEASYPRINT
from weasyprint import HTML

from django.core.paginator import Paginator

#OPENPYXL
from openpyxl import Workbook


# Create your views here.
def ListProveedores(request):
    context_provedores =  ProveedoresModels.objects.all()
    paginator = Paginator(context_provedores, 6)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number) 
    return render(request, 'index-proveedores.html', {'page_obj': page_obj})

def CreateProveedores(request):
    form =  ProveedorForm(request.POST) 
    if form.is_valid():
        form.save()
        return redirect('proveedores:list-proveedores')
    else:
        form =  ProveedorForm()
    return render(request, 'index-proveedores-form.html', {'form': form})

def UpdateProveedores(request, id):
    proveedor = get_object_or_404(ProveedoresModels, id=id)

    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            return redirect('proveedores:list-proveedores')
    else:
        form = ProveedorForm(instance=proveedor)

    return render(request, 'index-proveedores-form.html', {'form': form})

def DeleteProveedores(request, id):
    proveedor = get_object_or_404(ProveedoresModels, id=id)

    if request.method == 'POST':
        proveedor.state = False
        proveedor.save()
        return redirect('proveedores:list-proveedores')

    return render(request, 'index-proveedores-delete.html')

def ExportarProveedoresXML(request):
    wb = Workbook()
    ws = wb.active

    ws['B1'] = 'REPORTE DE PROVEEDORES'
    ws.merge_cells('B1:E1')

    ws['B2'] = 'NOMBRE' 
    ws['C2'] = 'TELEFONO'
    ws['D2'] = 'OCUPACION'
    ws['E2'] = 'PRECIO'

    count = 3
    for i in ProveedoresModels.objects.all():
        ws.cell(row=count, column=2).value = i.name_proveedores
        ws.cell(row=count, column=3).value = i.telephone
        ws.cell(row=count, column=4).value = i.ocupation
        ws.cell(row=count, column=5).value = i.price

        count += 1

    Reporte_Proveedores = "Crematorio_Reporte_Proveedores.xlsx"
    response = HttpResponse(content_type = "aplication/ms-excel")
    content = "attachment; filename = {0}".format(Reporte_Proveedores)
    response['Content-Disposition'] = content
    wb.save(response)

    return response


def ExportarProveedoresPDF(request):
    template =  get_template('report_proveedores_pdf.html')
    context =  {
        'prov': ProveedoresModels.objects.all()
    }
    html_render =  template.render(context)
    pdf = HTML(string=html_render).write_pdf()

    response =  HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="proveedores.pdf"'
    return response