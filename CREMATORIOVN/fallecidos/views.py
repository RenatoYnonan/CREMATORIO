from re import template
from django.shortcuts import render, redirect
from django.views.generic import *

from django.http import HttpResponse
from django.urls import reverse_lazy

from django.contrib.auth.mixins import LoginRequiredMixin

from django.core.paginator import Paginator

from django.template.loader import get_template

from weasyprint import HTML

#IMPORTACIONES 
from .models import *
from .forms import *

#COMPONENTES DE OPENPYXL
from openpyxl  import Workbook
from openpyxl.styles import Alignment

class ReporteFallecido(TemplateView):
    def get(self, request, *args, **kwargs):
        exfallecido =  Fallecido.objects.all()

        wb  = Workbook()
        ws = wb.active

        ws['B1'] = 'LISTADO DE FALLECIDOS'
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('B1:E1')

        ws['B3'] = 'NOMBRE'
        ws['C3'] = 'APELLIDO'
        ws['D3'] = 'FECHA DE NACIMIENTO'
        ws.column_dimensions['D'].width = 22
        ws['E3'] = 'FECHA DE FALLECIDO'
        ws.column_dimensions['E'].width = 22
        ws['F3'] = 'GENERO'
        ws['G3'] = 'DNI'
        
        contador = 4

        for i in exfallecido:
            ws.cell(row=contador, column=2).value = i.nombre_completo.capitalize()
            ws.cell(row=contador, column=3).value = i.apellido.capitalize()

            #FECHA DE NACIMIENTO
            ws.cell(row=contador, column=4).value = i.fecha_nacimiento
            ws.cell(row=contador, column=4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            #FECHA DE FALLECIDO
            ws.cell(row=contador, column=5).value = i.fecha_fallecido
            ws.cell(row=contador, column=5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


            ws.cell(row=contador, column=6).value = i.get_genero_display()
            ws.cell(row=contador, column=7).value = i.documento_identidad
        
            contador += 1

        Reporte_Fallecido = "Crematorio_Reporte_Fallecido.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(Reporte_Fallecido)
        response['Content-Disposition'] = content
        wb.save(response)

        return response

def ReporteFallecidoPDF(request):
    template = get_template('index_pdf.html')
    exfallecido =  {
        'fallecido': Fallecido.objects.all()
    }
    
    html = template.render(exfallecido)
    pdf = HTML(string=html).write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="fallecido.pdf"'

    return response


class Fallecidos(LoginRequiredMixin, ListView):
    model = Fallecido
    paginate_by = 6
    template_name = 'index_fallecido.html'


class CreateDeceased(LoginRequiredMixin, CreateView):
    model = Fallecido
    form_class = FallecidosForms
    template_name = 'index-fallecido-form.html'
    success_url =  reverse_lazy('fallecidos:fallecidos')


class UpdateDeceased(LoginRequiredMixin, UpdateView):
    model = Fallecido
    form_class = FallecidosForms
    template_name = 'index-fallecido-form.html'
    success_url =  reverse_lazy('fallecidos:fallecidos')

def DeleteDeceased(request, pk):
    fallecido = Fallecido.objects.get(id=pk)
    if request.method == 'POST':
        if fallecido.state == True:
            fallecido.state = False
        else:
            fallecido.state = True
        fallecido.save()
        return redirect('fallecidos:fallecidos')
    return render(request, 'index-delete.html')

    
    
