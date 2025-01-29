from django.shortcuts import render
from django.views.generic import *
from .models import *

from .forms import *

from django.http import HttpResponse, HttpResponseNotAllowed
from django.shortcuts import get_object_or_404, redirect

from django.template.loader import get_template

#OPENPYXL
from openpyxl import Workbook

#WEASYPRNT
from weasyprint import HTML

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin

# Create your views here.

class ListInstituciones(LoginRequiredMixin, ListView):
    model = ModelsInstitucion
    context_object_name = 'instituciones'
    paginate_by = 6
    template_name = 'instituciones/index-instituciones.html'

@login_required(login_url='intranet-crematorio:login')
def CreateInstitucion(request):

    if request.method == 'POST':
        form  =  InstitucionesForm(request.POST, request.FILES) 
        if form.is_valid():
            form.save()
            return redirect('convenios:list-instituciones')
        else:
            print(form.errors)
    else:
        form =  InstitucionesForm()
    return render(request, 'instituciones/index-instituciones-forms.html', {'form': form})

@login_required(login_url='intranet-crematorio:login')
def UpdateInstitucion(request, id):
    institucion =  get_object_or_404(ModelsInstitucion, id=id)

    if request.method == 'POST':
        form  =  InstitucionesForm(request.POST, request.FILES, instance=institucion)
        if form.is_valid():
            form.save()
            return redirect('convenios:list-instituciones')
        else:
            print(form.errors)
    else:
        form =  InstitucionesForm(instance=institucion)
    return render(request, 'instituciones/index-instituciones-forms.html', {'form': form})


@login_required(login_url='intranet-crematorio:login')
def DeleteInstitucion(request, id):
    institucion = get_object_or_404(ModelsInstitucion, id=id)   

    if request.method == 'POST':
        institucion.estado = False
        institucion.save()
        return redirect('convenios:list-instituciones')

    return render(request, 'instituciones/index-delete.html')

def ExportarInstitucionesXML(request):
    wb = Workbook()
    ws = wb.active

    ws['B1'] = 'REPORTE DE INSTITUCIONES'
    ws.merge_cells('B1:C1')

    ws['B2'] = 'NOMBRE INSTITUCION'
    ws['C2'] = 'CIUDAD'

    count = 3
    for i in ModelsInstitucion.objects.all():
        ws.cell(row=count, column=2).value = i.nombre_institucion
        ws.cell(row=count, column=3).value = i.ciudad

        count += 1

    Reporte_Instituciones = "Crematorio_Reporte_Instituciones.xlsx"
    response = HttpResponse(content_type = "aplication/ms-excel")
    content = "attachment; filename = {0}".format(Reporte_Instituciones)
    response['Content-Disposition'] = content
    wb.save(response)

    return response

def ExportarInstitucionesPDF(request):
    template =  get_template('instituciones/report_instituciones_pdf.html')
    context =  {
        'ins': ModelsInstitucion.objects.all()
    }
    html_render =  template.render(context)
    pdf = HTML(string=html_render).write_pdf()

    response =  HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'atachment; filename="instituciones.pdf"'
    return response

#FUNERARIAS

class ListFunerarias(LoginRequiredMixin, ListView):
    model = ModelsFunerarias
    paginate_by = 6
    context_object_name = 'funerarias'
    template_name = 'funerarias/index-funerarias.html'

@login_required(login_url='intranet-crematorio:login')
def CreateFuneraria(request):

    if request.method == 'POST':
        form  =  FuneriasForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('convenios:list-funerarias')
        else:
            print(form.errors)
    else:
        form =  FuneriasForm()
    return render(request, 'funerarias/index-funerarias-forms.html', {'form': form})

@login_required(login_url='intranet-crematorio:login')
def UpdateFuneraria(request, id):
    funeraria =  get_object_or_404(ModelsFunerarias, id=id)

    if request.method == 'POST':
        form  =  FuneriasForm(request.POST, request.FILES, instance=funeraria)
        if form.is_valid():
            form.save()
            return redirect('convenios:list-funerarias')
        else:
            print(form.errors)
    else:
        form =  FuneriasForm(instance=funeraria)
    return render(request, 'funerarias/index-funerarias-forms.html', {'form': form})

@login_required(login_url='intranet-crematorio:login')
def deleteFuneraria(request, id):
    funeraria = get_object_or_404(ModelsFunerarias, pk=id)

    if request.method == 'GET': 
        return render(request, 'funerarias/index-delete.html')

    if request.method == 'POST':  
        funeraria.estado = False  
        funeraria.save()
        return redirect('convenios:list-funerarias')
    
    return HttpResponseNotAllowed(['POST'])


def ExportarFunerariasXML(request):
    wb = Workbook()
    ws = wb.active

    ws['B1'] = 'REPORTE DE FUNERARIAS'
    ws.merge_cells('B1:C1')

    ws['B2'] = 'NOMBRE FUNERARIA'
    ws['C2'] = 'CIUDAD'

    count = 3
    for i in ModelsFunerarias.objects.all():
        ws.cell(row=count, column=2).value = i.nombre_funeraria
        ws.cell(row=count, column=3).value = i.ciudad

        count += 1

    Reporte_Funerarias = "Crematorio_Reporte_Funerarias.xlsx"
    response = HttpResponse(content_type = "aplication/ms-excel")
    content = "attachment; filename = {0}".format(Reporte_Funerarias)
    response['Content-Disposition'] = content
    wb.save(response)

    return response


def ExportarFunerariasPDF(request):
    template =  get_template('funerarias/report_funerarias_pdf.html')
    context =  {
        'fun': ModelsFunerarias.objects.all()
    }
    html_render =  template.render(context)
    pdf = HTML(string=html_render).write_pdf()

    response =  HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'atachment; filename="funerarias.pdf"'
    return response