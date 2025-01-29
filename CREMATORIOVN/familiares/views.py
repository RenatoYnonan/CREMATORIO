from os import error
from django.shortcuts import render
from django.views.generic import *
from .models import *
from django.urls import reverse_lazy
from .forms import *


from django.contrib.auth.decorators import login_required

from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from openpyxl import Workbook
from django.http import HttpResponseRedirect, HttpResponse

from django.contrib.auth.mixins import LoginRequiredMixin
from django.template.loader import get_template

#WEASYPRINT
from weasyprint import HTML

# Create your views here.
class ListFamily( LoginRequiredMixin,ListView):
    model = ModelsFamiliar
    paginate_by = 6
    template_name = 'index-familiar.html'

@login_required(login_url='intranet-crematorio:login')
def CreateFamiliar(request):
    if request.method == 'POST':
        form = FamiliaresForms(request.POST)
        if form.is_valid():
            form.save()
            return redirect('familiares:list-familiares')
    else:
        form = FamiliaresForms()
    return render(request, 'index-form-familiar.html', {'form': form})

class FamiliarUpdate(LoginRequiredMixin, UpdateView):
    model = ModelsFamiliar
    form_class = FamiliaresForms
    template_name = 'index-form-familiar.html'
    success_url =  reverse_lazy('familiares:list-familiares')


@login_required(login_url='intranet-crematorio:login')
def DeleteFamiliar(request, pk):
    familiares = get_object_or_404(ModelsFamiliar, pk=pk)
    if request.method == 'POST':
        if familiares.estado == True:
            familiares.estado = False
            familiares.save()
            return redirect('familiares:list-familiares')
        else:
            familiares.estado = True
            familiares.save()
            return redirect('familiares:list-familiares')
    else:
        return redirect('familiares:list-familiares')
    return render(request, 'index-delete-familiar.html')
    


class ReporteExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        exfami = ModelsFamiliar.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE FAMILIARES'

        ws.merge_cells('B1:E1')
        ws['B3'] = 'NOMBRES'
        ws['C3'] = 'APELLIDOS'

        contador = 2

        for i in exfami:
            ws.cell(row=contador, column=2).value = i.name.capitalize()
            ws.cell(row=contador, column=3).value = i.lastname.capitalize()

            contador += 1
        
        Reporte_Familiares = "Crematorio_Reporte_Familiares.xlsx"
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(Reporte_Familiares)
        response['Content-Disposition'] = content
        wb.save(response)

        return response

def ReportePDF(request):
    template =  get_template('report_familiares_pdf.html')
    context =  {
        'fam': ModelsFamiliar.objects.all()
    }
    html_render =  template.render(context)
    pdf = HTML(string=html_render).write_pdf()

    response =  HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="familiares.pdf"'
    return response