from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import *
from django.urls import reverse_lazy
from django.core.paginator import Paginator

from django.contrib import messages

from django.template.loader import get_template
from weasyprint import HTML

from django.http import HttpResponseRedirect, HttpResponse, response
from .models import *
from .forms  import *

from openpyxl import Workbook

from django.contrib.auth.decorators import login_required

# Create your views here.
@login_required(login_url='intranet-crematorio:login')
def ListProductos(request):
    product = ModelsProduct.objects.all()
    pag = Paginator(product, 6)
    page_number = request.GET.get('page')

    page_obj =  pag.get_page(page_number)


    return render(request, 'productos/index-productos.html', { 'page_obj': page_obj })


@login_required(login_url='intranet-crematorio:login')
def CreateProduct(request):
    if request.method == 'POST':
        form = FormsProduct(request.POST)

        if form.is_valid():
            form.save()
            return redirect('catalogo:list-product')
    else:
        form = FormsProduct()
    return render(request, 'productos/index-form.html', { 'form': form })


@login_required(login_url='intranet-crematorio:login')
def UpdateProduct(request, id):
    datos = get_object_or_404(ModelsProduct, id=id)

    if request.method == 'POST':
        form = FormsProduct(request.POST, instance=datos)

        if form.is_valid():
            form.save()
            return redirect('catalogo:list-product')
    else:
        form = FormsProduct(instance=datos)


    return render(request, 'productos/index-form.html', {'form': form})

@login_required(login_url='intranet-crematorio:login')
def DeleteProduct(request, id):
    data = get_object_or_404(ModelsProduct, id=id)

    if request.method == 'POST':
        data.state_product = False
        data.save()
        return redirect('catalogo:list-product')
    else:
        form = FormsProduct(instance=data)
    
    return render(request, 'productos/index-delete.html', {'form': form})

# GENERAR PDF PRODUCTOS
def exportar_pdf(request):

    html_string = get_template('productos/ticket.html')
    context = {
        'pds': ModelsProduct.objects.all()  
    } 

    html = html_string.render(context)
    pdf =  HTML(string=html).write_pdf()

    response =  HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="ticket.pdf"'
    return response

    #obtiene el template html de la factura
    #crea un contexto con todos los productos
    #renderiza el template con el contexto
    #convierte el html a pdf
    #crea una respuesta con el pdf
    #agrega un header para que el navegador lo abra directamente
    #devuelve la respuesta

def ExportarXML(request):
    ws = Workbook()
    wd = ws.active
    wd['B1'] = 'REPORTE DE PRODUCTOS'
    wd.merge_cells('B1:D1')

    wd['B2'] = 'NOMBRE DEL PRODUCTO'
    wd['C2'] = 'DESCRIPCION DEL PRODUCTO'
    wd['D2'] = 'PRECIO DEL PRODUCTO'

    count = 3

    for i in ModelsProduct.objects.all():
        wd.cell(row=count, column=2).value = i.name_product
        wd.cell(row=count, column=3).value = i.descriptions_product
        wd.cell(row=count, column=4).value = i.price_product

        count += 1

    response = HttpResponse(content_type = "aplication/ms-excel")
    content = "attachment; filename = {0}".format("Crematorio_Reporte_Productos.xlsx")
    response['Content-Disposition'] = content
    ws.save(response)

    return response




# PLANES DE CREMACION 
# - CREMACIÓN BASICA
# - CREMACION PREMIÚM 
# - PLAN A FUTURO  

@login_required(login_url='intranet-crematorio:login')
def ListPlanes(request): #listar datos
    obj_plan = ModelsPlanes.objects.all()

    pagination =  Paginator(obj_plan, 6)
    page_number = request.GET.get('page')
    page_obj =  pagination.get_page(page_number)

    return render(request, 'planes/index-planes.html', {'page_obj': page_obj})

@login_required(login_url='intranet-crematorio:login')
def CreatePlan(request):
    if request.method == 'POST':
        form_plan = FormPlanes(request.POST)

        if form_plan.is_valid():
            form_plan.save()
            return redirect('catalogo:list-planes')
    else: 
        form_plan = FormPlanes()


    return render(request, 'planes/index-form.html', {'form_plan': form_plan})

@login_required(login_url='intranet-crematorio:login')
def UpdatePlanes(request, id):
    
    obj_planes = get_object_or_404(ModelsPlanes, id=id)

    if request.method == 'POST':
        form_plan = FormPlanes(request.POST, instance=obj_planes)

        if form_plan.is_valid():
            form_plan.save()
            return redirect('catalogo:list-planes')
    else: 
        form_plan = FormPlanes(instance=obj_planes)


    return render(request, 'planes/index-form.html', {'form_plan': form_plan})


@login_required(login_url='intranet-crematorio:login')
def DeletePlanes(request, id):
    obj_planes = get_object_or_404(ModelsPlanes, id=id)

    if request.method == 'POST':

        if obj_planes.state_plan is True:
            obj_planes.state_plan = False
            obj_planes.save()
            return redirect('catalogo:list-planes')
        else:
            messages.success({f'El {obj_planes.name_plan} ya esta inactivo'})
    
    return render(request, 'planes/index-delete.html')


def ExportarPlanesXML(request):

    wd = Workbook()
    ws = wd.active

    ws['B1'] = 'REPORTE DE PLANES'
    ws.merge_cells('B1:C1')

    ws['B2'] = 'NOMBRE DEL PLAN'
    ws['C2'] = 'DESCRIPCION DEL PLAN'
    ws['D2'] = 'PRECIO DEL PLAN'

    count = 3
    for i in ModelsPlanes.objects.all():
        ws.cell(row=count, column=2).value = i.name_plan
        ws.cell(row=count, column=3).value = i.description_plan
        ws.cell(row=count, column=4).value = i.price_plan

        count += 1
    response = HttpResponse(content_type = "aplication/ms-excel")
    content = "attachment; filename = {0}".format("Crematorio_Reporte_Planes.xlsx")
    response['Content-Disposition'] = content
    wd.save(response)

    return response
    

def ExportarPlanesPdf(request):
    template_planes = get_template('planes/report_planes_pdf.html')
    context ={
        'planes': ModelsPlanes.objects.all()
    }
    html_render =  template_planes.render(context)
    template_pdf = HTML(string=html_render).write_pdf()

    response =  HttpResponse(template_pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'atachment; filename="planes.pdf"'
    return response


#EXPORTAR URNAS

class ListUrnas(ListView):
    model = ModelsUrnas
    template_name = 'urnas/index-urnas.html'
    paginate_by = 6

class CreateUrnas(CreateView):
    model = ModelsUrnas
    form_class = FormUrnas
    template_name = 'urnas/index-form.html'
    success_url = reverse_lazy('catalogo:list-urnas')

class UpdateUrnas(UpdateView):
    model = ModelsUrnas
    form_class = FormUrnas
    template_name = 'urnas/index-form.html'
    success_url = reverse_lazy('catalogo:list-urnas')
    slug_field = 'slug_urna'  
    slug_url_kwarg = 'slug_urna'

def DeleteUrnas(request, slug_urna):
    pass